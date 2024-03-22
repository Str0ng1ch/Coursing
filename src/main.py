import io
from collections import Counter
from datetime import datetime
from functools import wraps

import Levenshtein
import pandas as pd
import requests
from flask import Flask, render_template, request, jsonify, redirect, url_for, Response, session
from flask_mysqldb import MySQL
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.make_ratings import make_rating

# DATABASE, TABLE = "u2255198_coursing", "results_test"
DATABASE, TABLE = "coursing", "results"
ADMIN_USERNAME, ADMIN_PASSWORD = "tanya_admin", "p0n4ik"
VIEW_LEVREKTI_USERNAME, VIEW_LEVREKTI_PASSWORD = "view_access", "l3vr3tk1P@ssw0rd"
VIEW_PHARAOH_USERNAME, VIEW_PHARAOH_PASSWORD = "view_access", "password_phara0h_h0und"
SECRET_KEY = ")#lO4\\;nR<0Wy=y^CRM|{#;5f}1{Emu'zt]"

application = Flask(__name__)
# application.config['MYSQL_HOST'] = "server25.hosting.reg.ru"
# application.config['MYSQL_USER'] = "u2255198_artem"
# application.config['MYSQL_PASSWORD'] = "00zEbyTI3y5avEot"
application.config['MYSQL_HOST'] = "localhost"
application.config['MYSQL_USER'] = "root"
application.config['MYSQL_PASSWORD'] = "My$QLP@ssw0rd"
application.config['MYSQL_DB'] = DATABASE

mysql = MySQL(application)
application.secret_key = SECRET_KEY


def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization

        if auth and auth.username == ADMIN_USERNAME and auth.password == ADMIN_PASSWORD:
            return f(*args, **kwargs)

        return Response(
            'Could not verify your access level for that URL.\n'
            'You have to login with proper credentials.', 401,
            {'WWW-Authenticate': 'Basic realm="Login Required"'})

    return decorated


def requires_auth_view(f, username, password):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization

        if (auth and ((auth.username == username and auth.password == password)
                      or (auth.username == ADMIN_USERNAME and auth.password == ADMIN_PASSWORD))):
            return f(*args, **kwargs)

        return Response(
            'Could not verify your access level for that URL.\n'
            'You have to login with proper credentials.', 401,
            {'WWW-Authenticate': 'Basic realm="Login Required"'})

    return decorated


@application.route('/')
def index():
    return render_template('index.html')


@application.route('/racing')
def racing():
    return render_template('racing.html')


@application.route('/racing-rating/<param>')
def racing_rating(param):
    cursor = mysql.connection.cursor()
    query = f'SELECT DISTINCT(date) FROM {DATABASE}.racing'
    cursor.execute(query)
    dates = cursor.fetchall()
    cursor.close()
    if param == 'whippets':
        return render_template('racing_rating.html', date_options=[date[0].strftime("%Y-%m-%d") for date in dates])


@application.route('/rating/<param>')
def rating(param):
    if param == 'whippets':
        return render_template('rating.html', param=param)
    elif param == 'italian_greyhounds':
        return requires_auth_view(render_template, VIEW_LEVREKTI_USERNAME, VIEW_LEVREKTI_PASSWORD)('rating.html',
                                                                                                   param=param)
    elif param == 'pharaoh_hound':
        return requires_auth_view(render_template, VIEW_PHARAOH_USERNAME, VIEW_PHARAOH_PASSWORD)('rating.html',
                                                                                                 param=param)


@application.route('/best/<param>')
def best(param):
    if param == 'whippets':
        return render_template('best.html', param=param)
    elif param == 'italian_greyhounds':
        return requires_auth_view(render_template, VIEW_LEVREKTI_USERNAME, VIEW_LEVREKTI_PASSWORD)('best.html',
                                                                                                   param=param)
    elif param == 'pharaoh_hound':
        return requires_auth_view(render_template, VIEW_PHARAOH_USERNAME, VIEW_PHARAOH_PASSWORD)('best.html',
                                                                                                 param=param)


@application.route('/explanations')
def explanations():
    return render_template('explanations.html')


@application.route('/in-developing')
def developing():
    return render_template('developing.html')


def handle_bad_request(error):
    return render_template('errors/error400.html'), 400


def handle_unauthorized(error):
    return render_template('errors/error401.html'), 401


def handle_forbidden(error):
    return render_template('errors/error403.html'), 403


def handle_not_found(error):
    return render_template('errors/error404.html'), 404


def handle_not_allowed(error):
    return render_template('errors/error405.html'), 405


def handle_too_many_request(error):
    return render_template('errors/error429.html'), 429


def handle_internal_server(error):
    return render_template('errors/error500.html'), 500


def handle_service_unavailable(error):
    return render_template('errors/error503.html'), 503


def handle_gateway_timeout(error):
    return render_template('errors/error504.html'), 504


@application.route('/get-score-details', methods=['POST'])
def get_score_details():
    data = request.json
    name, dog_type = data['score'].split(', ')
    name = name.replace('<br>', '/')

    cursor = mysql.connection.cursor()
    query = f'SELECT date, position, max_position, score, link, Nickname, breedarchive_link, Sex, location FROM {DATABASE}.{TABLE} WHERE Nickname = "{name}" AND Type = "{dog_type}" ORDER BY date '
    cursor.execute(query)
    score_details = cursor.fetchall()
    cursor.close()

    return jsonify(score_details)


@application.route('/get-partial-data', methods=['POST'])
def get_partial_data():
    selected_sex = request.json['selectedSex']
    selected_type = request.json['selectedType']
    param = request.json['paramValue']

    if param == 'whippets':
        param = 'Уиппет'
    elif param == 'italian_greyhounds':
        param = 'Малая итальянская борзая (Левретка)'
    elif param == 'pharaoh_hound':
        param = 'Фараонова собака'

    cur = mysql.connection.cursor()
    base_query = f"""SELECT * FROM (SELECT Type, Sex, Nickname, breedarchive_link, SUM(Score) AS TotalScore, COUNT(*) AS RecordCount, Breed
                                    FROM {DATABASE}.{TABLE} 
                                    WHERE Date >= DATE_SUB(CURDATE(), INTERVAL 1 YEAR) AND Type != 'Юниоры' AND Breed = '{param}'
                                    GROUP BY Type, Sex, Nickname, breedarchive_link
                                    ) AS sub"""
    conditions = []
    params = []

    if selected_type != "all":
        conditions.append("Type=%s")
        params.append(selected_type)
    if selected_sex != "all":
        conditions.append("Sex=%s")
        params.append(selected_sex)

    if conditions:
        query = base_query + " WHERE " + " AND ".join(conditions)
    else:
        query = base_query

    query += " ORDER BY TotalScore DESC, RecordCount ASC, Nickname ASC"
    cur.execute(query, params)
    rows = cur.fetchall()
    data = [{"type": row[0], "sex": row[1], "name": row[2], "TotalScore": row[4], "breedLink": row[3],
             "RecordCount": row[5]} for row in
            rows]

    return jsonify(data)


@application.route('/set-results-date', methods=['POST'])
def set_results_date():
    session['RESULTS_DATE'] = request.json['value']
    return jsonify({"status": "success"})


@application.route('/get-results-data', methods=['POST'])
def get_results_data():
    param = request.json['paramValue']
    selected_type = request.json['selectedType']
    selected_sex = request.json['selectedSex']
    name_search = request.json['nameSearch']
    all_rows = request.json.get('allRows', False)
    date = session.get('RESULTS_DATE')
    if not date:
        date = '2022-05-21'

    if param == 'whippets':
        param = 'Уиппет'
    elif param == 'italian_greyhounds':
        param = 'Малая итальянская борзая (Левретка)'
    elif param == 'pharaoh_hound':
        param = 'Фараонова собака'

    cur = mysql.connection.cursor()
    base_query = f"""SELECT * FROM (SELECT Date, Location, Distance, Type, Sex, Nickname, title, time_1, 
                                    time_2, time_3, link, breedarchive_link, breed FROM {DATABASE}.racing 
                                    WHERE Type != 'Юниоры' AND Breed = '{param}' AND Date = '{date}'
                                    ) AS sub"""
    conditions = []
    params = []

    if selected_type != "all":
        conditions.append("Type=%s")
        params.append(selected_type)
    if selected_sex != "all":
        conditions.append("Sex=%s")
        params.append(selected_sex)
    if name_search:
        conditions.append("Nickname LIKE %s")
        params.append("%" + name_search + "%")

    if conditions:
        query = base_query + " WHERE " + " AND ".join(conditions)
    else:
        query = base_query

    query += " ORDER BY Title ASC"
    if not all_rows:
        query += " LIMIT 9"
    cur.execute(query, params)
    rows = cur.fetchall()

    data = [{"date": row[0], "location": row[1], "distance": row[2], "type": row[3], "sex": row[4],
             "nickname": row[5], "title": row[6], "time_1": row[7], "time_2": row[8], "time_3": row[9],
             "link": row[10], "breedarchive_link": row[11], "breed": row[12]} for row in
            rows]

    return jsonify(data)


@application.route('/get-all-data', methods=['POST'])
@requires_auth
def get_all_data():
    selected_id = request.json['selectedID']
    selected_date = request.json['selectedDate']
    selected_sex = request.json['selectedSex']
    name_search = request.json['nameSearch']
    selected_rating = request.json['selectedRating']
    selected_type = request.json['selectedType']
    selected_location = request.json['selectedLocation']
    selected_breed = request.json['selectedBreed']
    all_rows = request.json.get('allRows', False)

    if selected_breed == 'Уиппеты':
        selected_breed = 'Уиппет'
    elif selected_breed == 'Левретки':
        selected_breed = 'Малая итальянская борзая (Левретка)'
    elif selected_breed == 'Фараоны':
        selected_breed = 'Фараонова собака'

    cur = mysql.connection.cursor()
    base_query = f"""SELECT * FROM {DATABASE}.{TABLE}"""
    conditions = []
    params = []

    if selected_type != "all":
        conditions.append("Type=%s")
        params.append(selected_type)
    if selected_sex != "all":
        conditions.append("Sex=%s")
        params.append(selected_sex)
    if selected_location != "all":
        conditions.append("Location=%s")
        params.append(selected_location)
    if selected_breed != "all":
        conditions.append("Breed=%s")
        params.append(selected_breed)
    if name_search:
        conditions.append("Nickname LIKE %s")
        params.append("%" + name_search + "%")
    if selected_rating != "all":
        rating_range = selected_rating.split('-')
        conditions.append("Score BETWEEN %s AND %s")
        params.extend([rating_range[0], rating_range[1]])
    if selected_date != "":
        conditions.append("Date=%s")
        params.append(selected_date)
    if selected_id != '':
        conditions.append("ID LIKE %s")
        params.append("%" + selected_id + "%")

    if conditions:
        query = base_query + " WHERE " + " AND ".join(conditions)
    else:
        query = base_query

    query += " ORDER BY Date DESC, Nickname ASC"
    if not all_rows:
        query += " LIMIT 9"

    cur.execute(query, params)
    rows = cur.fetchall()
    data = [{"id": row[0], "date": row[1], "position": row[2], "type": row[3], "sex": row[4], "Nickname": row[5],
             "max_position": row[6], "score": row[7], "link": row[8], "breedLink": row[9],
             "ignored": row[10], "breed": row[11], "location": row[12]} for row in rows]

    return jsonify(data)


def delete_from_table():
    message = None
    cursor = mysql.connection.cursor()
    try:
        cursor.execute(f"DROP TABLE IF EXISTS {DATABASE}.{TABLE}_copy; "
                       f"CREATE TABLE {DATABASE}.{TABLE}_copy AS SELECT * FROM {DATABASE}.{TABLE};")
        cursor.execute(f"DELETE FROM {DATABASE}.{TABLE};")
        cursor.execute(f"ALTER TABLE {DATABASE}.{TABLE} AUTO_INCREMENT = 1;")
        mysql.connection.commit()
        message = "Все данные успешно удалены!"
    except Exception as e:
        mysql.connection.rollback()
        message = f"Ошибка при удалении данных! \nОшибка: {e}"
    finally:
        cursor.close()
        return message


def add_data_from_excel():
    file = request.files['excel_file']
    cursor = mysql.connection.cursor()
    message = None

    try:
        df = pd.read_excel(file)

        values_list = df.apply(lambda row: (
            row['Дата'], row['Место'], row['Класс'], row['Пол'], row['Кличка'], row['Всего мест'],
            row['Очки'], row['Ссылка на источник'], row['Ссылка на Breed Archive'], row['Игнорированные ошибки'],
            row['Место соревнования'], row['Порода']
        ), axis=1)
        values_list = values_list.apply(lambda tpl: tuple('' if pd.isna(val) else val for val in tpl)).tolist()

        query = (f"INSERT INTO {DATABASE}.{TABLE} "
                 f"(Date, Position, Type, Sex, Nickname, max_position, Score, link, breedarchive_link, ignored, Location, Breed)"
                 f" VALUES (%s, %s, %s, %s, UPPER(%s), %s, %s, %s, %s, %s, %s, %s)")

        cursor.executemany(query, values_list)
        mysql.connection.commit()
        message = "Данные успешно внесены из Excel файла!"
    except Exception as e:
        mysql.connection.rollback()
        message = f"Ошибка при занесении данных из Excel файла! \nОшибка: {e}"
    finally:
        cursor.close()
        return message


def run_script():
    link = request.form.get('link', '')
    breed = request.form.get('breed', '')

    if breed == 'all':
        breed_list = ['Уиппет', 'Малая итальянская борзая (Левретка)', 'Фараонова собака']
    elif breed == 'Уиппеты':
        breed_list = ['Уиппет']
    elif breed == 'Левретки':
        breed_list = ['Малая итальянская борзая (Левретка)']
    else:
        breed_list = ['Фараонова собака']

    message = None
    try:
        make_rating(link, breed_list)
        message = f"Данные успешно внесены из {link}"
    except Exception as e:
        message = f"Ошибка при занесении данных из {link}. \nОшибка: {e}"
    finally:
        return message


def calculate_language(text):
    russian_letters = "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"
    english_letters = "abcdefghijklmnopqrstuvwxyz"

    russian_count, english_count = 0, 0
    text = text.lower()

    for char in text:
        if char in russian_letters:
            russian_count += 1
        elif char in english_letters:
            english_count += 1

    if russian_count == 0 and english_count == 0:
        return "na"
    if russian_count >= english_count:
        return "ru"
    return "en"


def update_nickname(nickname, row_id, cursor):
    query = f'UPDATE {DATABASE}.{TABLE} SET Nickname = "{nickname}" WHERE ID = {row_id}'
    cursor.execute(query)
    mysql.connection.commit()


def find_most_frequent_nickname(nicknames):
    filtered_names = [name for name in nicknames if name != ""]
    if len(filtered_names) > 0:
        name_counts = Counter(filtered_names)
        most_common_name, _ = name_counts.most_common(1)[0]

        return most_common_name
    return ''


def change_parts(dataset, cursor):
    for i in range(len(dataset)):
        update_nickname(dataset.iloc[i]["Nickname"].strip(), dataset.iloc[i]["ID"], cursor)
        if '/' not in dataset.iloc[i]['Nickname']:
            if calculate_language(dataset.iloc[i]['Nickname']) == 'ru':
                update_nickname("/" + dataset.iloc[i]["Nickname"], dataset.iloc[i]["ID"], cursor)

            elif calculate_language(dataset.iloc[i]['Nickname']) == 'en':
                update_nickname(dataset.iloc[i]["Nickname"] + "/", dataset.iloc[i]["ID"], cursor)
        else:
            nicknames = dataset.iloc[i]['Nickname'].split("/")
            if calculate_language(nicknames[0]) == 'ru':
                update_nickname(nicknames[1] + "/" + nicknames[0], dataset.iloc[i]["ID"], cursor)
            if calculate_language(nicknames[1]) == 'en':
                update_nickname(nicknames[1] + "/" + nicknames[0], dataset.iloc[i]["ID"], cursor)


def autofill_data():
    message = None
    cursor = mysql.connection.cursor()
    try:
        query = f"SELECT * FROM {DATABASE}.{TABLE}"
        cursor.execute(query)
        column_names = [desc[0] for desc in cursor.description]
        dataset = pd.DataFrame(cursor.fetchall(), columns=column_names)

        change_parts(dataset, cursor)

        query = f"SELECT * FROM {DATABASE}.{TABLE}"
        cursor.execute(query)
        column_names = [desc[0] for desc in cursor.description]
        dataset = pd.DataFrame(cursor.fetchall(), columns=column_names)
        dataset[['EnNickname', 'RuNickname']] = dataset['Nickname'].str.split('/', n=1, expand=True)

        for i in range(len(dataset)):
            current_dataset_en = dataset[
                (dataset["EnNickname"] == dataset.iloc[i]["EnNickname"]) & (dataset["EnNickname"] != "")]
            en_ids = current_dataset_en["ID"].tolist()
            en_nicknames = current_dataset_en["RuNickname"].tolist()

            current_dataset_ru = dataset[
                (dataset["RuNickname"] == dataset.iloc[i]["RuNickname"]) & (dataset["RuNickname"] != "")]
            ru_ids = current_dataset_ru["ID"].tolist()
            ru_nicknames = current_dataset_ru["EnNickname"].tolist()

            most_common_name_en = find_most_frequent_nickname(en_nicknames)
            most_common_name_ru = find_most_frequent_nickname(ru_nicknames)

            for j in range(len(en_ids)):
                if dataset[dataset['ID'] == en_ids[j]]["RuNickname"].item() != most_common_name_en:
                    update_nickname(
                        dataset[dataset['ID'] == en_ids[j]]["EnNickname"].item() + "/" + most_common_name_en,
                        dataset[dataset['ID'] == en_ids[j]]['ID'].item(), cursor)
            for j in range(len(ru_ids)):
                if dataset[dataset['ID'] == ru_ids[j]]["EnNickname"].item() != most_common_name_ru:
                    update_nickname(
                        dataset[dataset['ID'] == ru_ids[j]]["RuNickname"].item() + "/" + most_common_name_ru,
                        dataset[dataset['ID'] == ru_ids[j]]['ID'].item(), cursor)

        query = f"SELECT * FROM {DATABASE}.{TABLE}"
        cursor.execute(query)
        column_names = [desc[0] for desc in cursor.description]
        dataset = pd.DataFrame(cursor.fetchall(), columns=column_names)
        change_parts(dataset, cursor)

        query = f"""UPDATE {DATABASE}.{TABLE} AS t1
            JOIN (
                SELECT Nickname, MAX(breedarchive_link) AS breedarchive_link
                FROM {DATABASE}.{TABLE}
                WHERE breedarchive_link IS NOT NULL
                GROUP BY Nickname
                HAVING COUNT(Nickname) > 1
            ) AS t2 ON t1.Nickname = t2.Nickname
            SET t1.breedarchive_link = t2.breedarchive_link;
            """
        cursor.execute(query)
        mysql.connection.commit()

        query = f"""UPDATE {DATABASE}.{TABLE} SET Score = 0 WHERE Position = 0;"""
        cursor.execute(query)
        mysql.connection.commit()
        message = f"Данные успешно заполнены"
    except Exception as e:
        mysql.connection.rollback()
        message = f"Ошибка при заполнении данных. \nОшибка: {e}"
    finally:
        cursor.close()
        return message


def reset_ignored():
    message = None
    cursor = mysql.connection.cursor()
    try:
        cursor.execute(f"UPDATE {DATABASE}.{TABLE} SET ignored = ''")
        mysql.connection.commit()
        message = "Ошибки успешно сброшены"
    except Exception as e:
        mysql.connection.rollback()
        message = f"Ошибка при сбросе ошибок. \nОшибка: {e}"
    finally:
        cursor.close()
        return message


def add_data_from_form():
    message = None
    date = request.form['date']
    position = request.form['position']
    dog_type = request.form['type']
    sex = request.form['sex']
    nickname = request.form['nickname']
    max_position = request.form['max_position']
    score = request.form['score']
    link = request.form['link']
    breed_link = request.form['breed_link']
    location = request.form['location']
    breed = request.form['breed']

    if breed == 'Уиппеты':
        breed = 'Уиппет'
    elif breed == 'Левретки':
        breed = 'Малая итальянская борзая (Левретка)'
    elif breed == 'Фараоны':
        breed = 'Фараонова собака'

    cur = mysql.connection.cursor()
    try:
        cur.execute(
            f"INSERT INTO {DATABASE}.{TABLE} "
            f"(Date, Position, Type, Sex, Nickname, max_position, Score, link, breedarchive_link, breed, location) "
            "VALUES (%s, %s, %s, %s, UPPER(%s), %s, %s, %s, %s, %s, %s)",
            (date, position, dog_type, sex, nickname, max_position, score, link, breed_link, breed, location))
        mysql.connection.commit()
        message = "Данные успешно внесены из формы!"
    except Exception as e:
        mysql.connection.rollback()
        message = f"Ошибка при занесении данных из формы! \nОшибка: {e}"
    finally:
        cur.close()
        return message


def download_excel(cursor, full=True):
    query = f"SELECT * FROM {DATABASE}.{TABLE}"
    cursor.execute(query)
    result = cursor.fetchall()

    column_names = ['ID', 'Дата', 'Место', 'Класс', 'Пол', 'Кличка',
                    'Всего мест', 'Очки', 'Ссылка на источник', 'Ссылка на Breed Archive',
                    'Игнорированные ошибки', 'Порода', 'Место соревнования']

    wb = Workbook()
    sheet = wb.active

    sheet.append(column_names)
    if full:
        for row in result:
            sheet.append(row)

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if isinstance(cell.value, datetime):
                max_length = max(max_length, len(cell.value.strftime("%Y-%m-%d")))
            else:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    excel_data = io.BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    return excel_data


@application.route('/add-data', methods=['GET', 'POST'])
@requires_auth
def add_data():
    message = session.pop('message', None)
    if request.method == 'POST':
        if 'delete_button' in request.form:
            message = delete_from_table()
        elif 'excel_file' in request.files:
            message = add_data_from_excel()
        elif 'run_script' in request.form:
            message = run_script()
        elif 'reset_ignored' in request.form:
            message = reset_ignored()
        elif 'autofill_data' in request.form:
            message = autofill_data()
        else:
            message = add_data_from_form()
        session['message'] = message
        return redirect(url_for('add_data'))
    if 'download_full_excel' in request.args:
        cursor = mysql.connection.cursor()
        try:
            excel_data = download_excel(cursor, full=True)
            return Response(
                excel_data.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment;filename=database.xlsx"}
            )
        except Exception as e:
            message = f"Ошибка при скачивании Excel файла! \nОшибка: {e}"
        finally:
            cursor.close()
    elif 'download_part_excel' in request.args:
        cursor = mysql.connection.cursor()
        try:
            excel_data = download_excel(cursor, full=False)
            return Response(
                excel_data.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment;filename=template.xlsx"}
            )
        except Exception as e:
            message = f"Ошибка при скачивании Excel файла! \nОшибка: {e}"
        finally:
            cursor.close()

    return render_template('add_data.html', message=message)


@application.route("/update-data", methods=["POST"])
@requires_auth
def update_data():
    updated_data = request.json
    row_id = updated_data['id']
    date = updated_data['date'].split('.')
    position = updated_data['position']
    dog_type = updated_data['type']
    sex = updated_data['sex']
    nickname = updated_data['nickname']
    max_position = updated_data['max_position']
    score = updated_data['score']
    link = updated_data['link']
    breed_link = updated_data['BreedLink']
    breed = updated_data['breed']
    location = updated_data['location']

    formatted_date = f"{date[2]}-{date[1]}-{date[0]}"
    cur = mysql.connection.cursor()
    try:
        update_query = f"""
            UPDATE {DATABASE}.{TABLE}
            SET
              Date = %s,
              Position = %s,
              Type = %s,
              Sex = %s,
              Nickname = UPPER(%s),
              max_position = %s,
              Score = %s,
              link = %s,
              breedarchive_link = %s,
              breed = %s,
              location = %s
            WHERE
              ID = %s
            """
        cur.execute(update_query,
                    (formatted_date, position, dog_type, sex, nickname, max_position, score, link, breed_link, breed,
                     location, row_id))
        mysql.connection.commit()
        return jsonify()
    except Exception:
        mysql.connection.rollback()
        return jsonify(), 500
    finally:
        cur.close()


def check_url_status(url):
    try:
        response = requests.get(url)
        return response.status_code
    except requests.exceptions.RequestException:
        return None


def find_similar_pairs_with_distance(values_list, min_distance, max_distance):
    similar_pairs = []
    for i in range(len(values_list)):
        for j in range(i + 1, len(values_list)):
            distance = Levenshtein.distance(values_list[i], values_list[j])
            if min_distance <= distance <= max_distance:
                similar_pairs.append((values_list[i], values_list[j]))

    return similar_pairs


@application.route("/check-database-without-links", methods=["POST"])
@requires_auth
def check_database_without_links():
    return check_data(check_links=False)


@application.route("/check-database-with-links", methods=["POST"])
@requires_auth
def check_database_with_links():
    return check_data()


def check_data(check_links=True):
    cursor = mysql.connection.cursor()
    cursor.execute(f"SELECT * FROM {DATABASE}.{TABLE}")

    ids, errors = [], []
    column_names = [desc[0] for desc in cursor.description]
    dataset = pd.DataFrame(cursor.fetchall(), columns=column_names)

    mask = (~dataset['Sex'].isin(['Кобель', 'Сука'])) & ~dataset['ignored'].astype(str).str.contains('1')
    result = dataset[mask]
    ids.extend(result['ID'].to_list())
    errors.extend(['Неправильно указан пол'] * len(result['ID'].to_list()))

    mask = (~dataset['Type'].isin(['Стандартный', 'Стандартный-спринтеры', 'Юниоры'])) & ~(
        dataset['ignored'].astype(str).str.contains('2'))
    result = dataset[mask]
    ids.extend(result['ID'].to_list())
    errors.extend(['Неправильно указан класс'] * len(result['ID'].to_list()))

    mask = (dataset['Score'] != dataset['Max_position'] - dataset['Position'] + 1) & ~(
        dataset['ignored'].astype(str).str.contains('3')) & ~(
            (dataset['Position'] == 0) & (dataset['Score'] == 0))
    result = dataset[mask]
    ids.extend(result['ID'].to_list())
    errors.extend(['Неправильно посчитаны очки'] * len(result['ID'].to_list()))

    if check_links:
        unique_links = dataset['link'].unique()
        link_status_dict = {link: check_url_status(link) for link in unique_links}
        dataset['Link_status'] = dataset['link'].map(link_status_dict)

        unique_breed_links = dataset['breedarchive_link'].unique()
        breed_link_status_dict = {link: check_url_status(link) for link in unique_breed_links}
        dataset['breedarchive_link_status'] = dataset['breedarchive_link'].map(breed_link_status_dict)

        result = dataset[(dataset['Link_status'] != 200) & ~(dataset['ignored'].astype(str).str.contains('4'))]
        ids.extend(result['ID'].to_list())
        errors.extend(['Недействительная ссылка на результаты соревнований'] * len(result['ID'].to_list()))

        result = dataset[
            (dataset['breedarchive_link_status'] != 200) & ~(dataset['ignored'].astype(str).str.contains('5'))]
        ids.extend(result['ID'].to_list())
        errors.extend(['Недействительная ссылка на BreedArchive'] * len(result['ID'].to_list()))

    result = dataset[
        (dataset['Nickname'].str.contains('/') == False) & ~(dataset['ignored'].astype(str).str.contains('-'))]
    ids.extend(result['ID'].to_list())
    errors.extend(['Нет разделения на между английской или русской частью'] * len(result['ID'].to_list()))

    try:
        dataset[['Part1', 'Part2']] = dataset['Nickname'].str.split('/', n=1, expand=True)
        similar_pairs = find_similar_pairs_with_distance(dataset['Part1'], 1, 2)
        for i in range(len(similar_pairs)):
            first_name, second_name = similar_pairs[i]
            result = dataset[(dataset['Part1'] == first_name) & ~(dataset['ignored'].astype(str).str.contains('#'))]
            ids.extend(result['ID'].to_list())
            errors.extend(['Возможна опечатка в русском имени'] * len(result['ID'].to_list()))
            result = dataset[(dataset['Part1'] == second_name) & ~(dataset['ignored'].astype(str).str.contains('#'))]
            ids.extend(result['ID'].to_list())
            errors.extend(['Возможна опечатка в русском имени'] * len(result['ID'].to_list()))

        similar_pairs = find_similar_pairs_with_distance(dataset['Part2'], 1, 2)
        for i in range(len(similar_pairs)):
            first_name, second_name = similar_pairs[i]
            result = dataset[(dataset['Part2'] == first_name) & ~(dataset['ignored'].astype(str).str.contains('@'))]
            ids.extend(result['ID'].to_list())
            errors.extend(['Возможна опечатка в английском имени'] * len(result['ID'].to_list()))
            result = dataset[(dataset['Part2'] == second_name) & ~(dataset['ignored'].astype(str).str.contains('@'))]
            ids.extend(result['ID'].to_list())
            errors.extend(['Возможна опечатка в английском имени'] * len(result['ID'].to_list()))

        grouped = dataset[dataset['Part2'] != ''].groupby('Part2')['ID'].unique().reset_index()
        for row in grouped[grouped['ID'].apply(len) >= 2].values:
            subset = dataset[dataset['ID'].isin(row[1])]
            if subset['Part1'].nunique() != 1:
                for row_id in row[1]:
                    result = dataset[(dataset['ID'] == row_id) & (~dataset['ignored'].astype(str).str.contains('&'))]
                    ids.extend(result['ID'].to_list())
                    errors.extend(
                        [f'Не совпадает английское написание имени среди собак с такой же русском кличкой'] * len(
                            result['ID'].to_list()))

        grouped = dataset[dataset['Part1'] != ''].groupby('Part1')['ID'].unique().reset_index()
        for row in grouped[grouped['ID'].apply(len) >= 2].values:
            subset = dataset[dataset['ID'].isin(row[1])]
            if subset['Part2'].nunique() != 1:
                for row_id in row[1]:
                    result = dataset[(dataset['ID'] == row_id) & ~(dataset['ignored'].astype(str).str.contains('%'))]
                    ids.extend(result['ID'].to_list())
                    errors.extend(
                        [f'Не совпадает русское написание имени среди собак с такой же английской кличкой'] * len(
                            result['ID'].to_list()))

        result = dataset[(dataset['Part1'] == "") & ~(dataset['ignored'].astype(str).str.contains('9'))]
        ids.extend(result['ID'].to_list())
        errors.extend(["Не заполнена английская часть клички"] * len(result['ID'].to_list()))

        result = dataset[(dataset['Part2'] == "") & ~(dataset['ignored'].astype(str).str.contains('!'))]
        ids.extend(result['ID'].to_list())
        errors.extend(["Не заполнена русская часть клички"] * len(result['ID'].to_list()))
    except Exception:
        pass

    today = datetime.now().date()
    mask = (dataset['Date'] <= today)
    result = dataset.loc[(~mask) & ~(dataset['ignored'].astype(str).str.contains('6'))]
    ids.extend(result['ID'].to_list())
    errors.extend(['Недействительная дата'] * len(result['ID'].to_list()))

    mask = (dataset['Position'] > dataset['Max_position']) & ~(dataset['ignored'].astype(str).str.contains('7'))
    result = dataset[mask]
    ids.extend(result['ID'].to_list())
    errors.extend(['Позиция больше количества участников'] * len(result['ID'].to_list()))

    mask = (dataset['Score'] < 0) & ~(dataset['ignored'].astype(str).str.contains('8'))
    result = dataset[mask]
    ids.extend(result['ID'].to_list())
    errors.extend(['Количество очков меньше 0'] * len(result['ID'].to_list()))

    return jsonify([{"id": row[0], "error": row[1]} for row in zip(ids, errors)])


@application.route("/ignore-data", methods=["POST"])
def ignore_data():
    errors = {
        "Неправильно указан пол": '1',
        'Неправильно указан класс': '2',
        'Неправильно посчитаны очки': '3',
        'Недействительная ссылка на результаты соревнований': '4',
        'Недействительная ссылка на BreedArchive': '5',
        'Недействительная дата': '6',
        'Позиция больше количества участников': '7',
        'Количество очков меньше 0': '8',
        'Не заполнена английская часть клички': '9',
        'Не заполнена русская часть клички': '!',
        'Возможна опечатка в английском имени': '@',
        'Возможна опечатка в русском имени': '#',
        'Не совпадает английское написание имени среди собак с такой же русском кличкой': '&',
        'Не совпадает русское написание имени среди собак с такой же английской кличкой': '%',
        'Нет разделения на между английской или русской частью': '-'
    }
    row_id = request.json['id']
    error = request.json['error']
    cursor = mysql.connection.cursor()

    try:
        query = f"UPDATE {DATABASE}.{TABLE} SET ignored = CONCAT(ignored, '{errors[error]}') WHERE ID = {row_id}"
        cursor.execute(query)
        mysql.connection.commit()
        return jsonify()
    except Exception:
        mysql.connection.rollback()
        return jsonify(), 500
    finally:
        cursor.close()


def get_score_details_sections(breed, dog_type, sex):
    if breed == 'whippets':
        breed = 'Уиппет'
    elif breed == 'italian_greyhounds':
        breed = 'Малая итальянская борзая (Левретка)'
    elif breed == 'pharaoh_hound':
        breed = 'Фараонова собака'
    cursor = mysql.connection.cursor()
    query = f'''SELECT Nickname, TotalScore, breedarchive_link, Type, Sex, RecordCount FROM (SELECT Type, Sex, Nickname, breedarchive_link, SUM(Score) AS TotalScore, COUNT(*) AS RecordCount, Breed
                                    FROM {DATABASE}.{TABLE} 
                                    WHERE Date >= DATE_SUB(CURDATE(), INTERVAL 1 YEAR) AND Type != 'Юниоры' AND Breed = '{breed}'
                                    GROUP BY Type, Sex, Nickname, breedarchive_link
                                    ) AS sub WHERE Type = "{dog_type}" AND Sex = "{sex}" ORDER BY TotalScore DESC, RecordCount ASC, Nickname ASC LIMIT 5'''
    cursor.execute(query)
    score_details = cursor.fetchall()

    cursor.close()

    return jsonify(score_details)


@application.route("/get-score-details-section1", methods=["GET"])
def get_score_details_section1():
    breed = request.args.get('breed', type=str)
    return get_score_details_sections(breed, "Стандартный", "Кобель")


@application.route("/get-score-details-section2", methods=["GET"])
def get_score_details_section2():
    breed = request.args.get('breed', type=str)
    return get_score_details_sections(breed, "Стандартный", "Сука")


@application.route("/get-score-details-section3", methods=["GET"])
def get_score_details_section3():
    breed = request.args.get('breed', type=str)
    return get_score_details_sections(breed, "Стандартный-спринтеры", "Кобель")


@application.route("/get-score-details-section4", methods=["GET"])
def get_score_details_section4():
    breed = request.args.get('breed', type=str)
    return get_score_details_sections(breed, "Стандартный-спринтеры", "Сука")


error_handlers = {
    400: handle_bad_request,
    401: handle_unauthorized,
    403: handle_forbidden,
    404: handle_not_found,
    405: handle_not_allowed,
    429: handle_too_many_request,
    500: handle_internal_server,
    503: handle_service_unavailable,
    504: handle_gateway_timeout
}

for error_code, handler in error_handlers.items():
    application.register_error_handler(error_code, handler)


def create_app():
    return application


if __name__ == '__main__':
    application.run(host='0.0.0.0')
